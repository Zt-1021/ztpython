# 读取内容表格，表格首行为字段名（字段名及其顺序未知）。从第二行起均为表格实际内容
# 输入：表格的文件位置
# 输出：表格内容，以列表形式呈现，列表内每个元素为一字典，字典内部对应该列的字段名及相应值

from openpyxl import load_workbook


# wb = load_workbook("test_24_excel.xlsx")
# sheet = wb['test_data']
#
# keyName01 = sheet.cell(1, 1).value
# keyName02 = sheet.cell(1, 2).value
# keyName03 = sheet.cell(1, 3).value
#
# datas = []
# for i in range(2, sheet.max_row+1):
#     data = {}
#     data[keyName01] = sheet.cell(i, 1).value
#     data[keyName02] = sheet.cell(i, 2).value
#     data[keyName03] = sheet.cell(i, 3).value
#     datas.append(data)
# print(datas)

class ExcelData:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname
        wb = load_workbook(self.filename)
        global sheet
        sheet = wb[self.sheetname]

    def read_data(self):
        datas = []
        for i in range(2, sheet.max_row+1):
            data = {}
            data[sheet.cell(1, 1).value] = sheet.cell(i, 1).value
            data[sheet.cell(1, 2).value] = sheet.cell(i, 2).value
            data[sheet.cell(1, 3).value] = sheet.cell(i, 3).value
            datas.append(data)
        return datas


if __name__ == "__main__":
    data = ExcelData("test_24_excel.xlsx", "test_data").read_data()
    print(data)
