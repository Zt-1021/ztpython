"""学习excel的操作"""
import openpyxl


# 创建excel文件
# wb = openpyxl.Workbook()
# wb.save("test_data.xlsx")

# 定位excel文件
wb = openpyxl.load_workbook("test_data.xlsx")
# 定位表单
sheet = wb['Sheet']
# 表单存储值
# sheet.cell(1, 1).value = "case_id"
# 读取表单值
data = sheet.cell(3, 5).value
print("读取的值为{}".format(data))
# 保存表单
# wb.save("test_data.xlsx")
