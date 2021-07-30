"""excel类--测试数据在excel中"""
import openpyxl
from study.test.MathMethod import MathMethod


class DoExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_test_data(self):
        cases = []
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb[self.sheetname]
        for i in range(2, sheet.max_row+1):
            case = []
            case.append(sheet.cell(i, 1).value)  # case_id
            case.append(sheet.cell(i, 2).value)  # case_name
            case.append(sheet.cell(i, 3).value)  # a
            case.append(sheet.cell(i, 4).value)  # b
            case.append(sheet.cell(i, 5).value)  # expected
            cases.append(case)
        return cases

    def write_test_data(self, case_id, result, test_result):
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb[self.sheetname]
        sheet.cell(case_id+1, 6).value = result
        sheet.cell(case_id+1, 7).value = test_result
        wb.save(self.filename)
