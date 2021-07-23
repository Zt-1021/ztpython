import openpyxl
import common.constant


class DoExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_test_data(self):
        cases = []
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb[self.sheetname]
        for i in range(2, sheet.max_row+1):
            case = []
            case.append(sheet.cell(i, 1).value)  # id
            case.append(sheet.cell(i, 2).value)  # case_name
            case.append(sheet.cell(i, 3).value)  # url
            case.append(sheet.cell(i, 4).value)  # method
            case.append(sheet.cell(i, 5).value)  # data
            case.append(sheet.cell(i, 6).value)  # excepted
            cases.append(case)
        return cases

    def write_test_data(self, case_id, result, test_result):
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb[self.sheetname]
        sheet.cell(case_id+1, 7).value = result
        sheet.cell(case_id+1, 8).value = test_result
        wb.save(self.filename)


if __name__ == "__main__":

    path = common.constant.case_dir_admin
    # data_path = "/".join(path.split("\\"))
    case = DoExcel(path, "companybasicinfo").read_test_data()
    print(case)
