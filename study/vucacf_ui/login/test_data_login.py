import openpyxl


class DoExcel:
    def read_data(self):
        cases = []
        wb = openpyxl.load_workbook("test_data.xlsx")
        sheet = wb['login']
        for i in range(2, sheet.max_row+1):
            case = []
            case_id = sheet.cell(i, 1).value
            case_name = sheet.cell(i, 2).value
            username = sheet.cell(i, 3).value
            password = sheet.cell(i, 4).value
            vercode = sheet.cell(i, 5).value
            expected = sheet.cell(i, 6).value

            case.append(case_id)
            case.append(case_name)
            case.append(username)
            case.append(password)
            case.append(vercode)
            case.append(expected)

            cases.append(case)
        return cases
